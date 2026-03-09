"""
Generate a multi-department, multi-sub-department Excel file
in the same format as Dummy 3.xlsx.

Departments & Sub-departments structure:
- Engineering
  - B.TECH-INFORMATION TECHNOLOGY
  - B.TECH-COMPUTER SCIENCE AND ENGINEERING
  - B.TECH-ELECTRONICS AND COMMUNICATION ENGINEERING
- Postgraduate Engineering
  - M.E-COMPUTER SCIENCE AND ENGINEERING
  - M.E-VLSI DESIGN
- Management
  - MBA-BUSINESS ADMINISTRATION
  - MBA-FINANCE

Each program has students with multiple subjects.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random

# ── Data definitions ────────────────────────────────────────────────────────

DEPARTMENTS = {
    "Engineering": [
        "B.TECH-INFORMATION TECHNOLOGY",
        "B.TECH-COMPUTER SCIENCE AND ENGINEERING",
        "B.TECH-ELECTRONICS AND COMMUNICATION ENGINEERING",
    ],
    "Postgraduate Engineering": [
        "M.E-COMPUTER SCIENCE AND ENGINEERING",
        "M.E-VLSI DESIGN",
    ],
    "Management": [
        "MBA-BUSINESS ADMINISTRATION",
        "MBA-FINANCE",
    ],
}

# Subjects per program  {program: [(code, name, credits), ...]}
SUBJECTS = {
    "B.TECH-INFORMATION TECHNOLOGY": [
        ("U18ITI7203L-R21", "MACHINE LEARNING", 3),
        ("U18ITI7201L-R21", "DATA STRUCTURES", 4),
        ("U18ITI7202L-R21", "COMPUTER NETWORKS", 3),
    ],
    "B.TECH-COMPUTER SCIENCE AND ENGINEERING": [
        ("U18CSI7301L-R21", "OPERATING SYSTEMS", 4),
        ("U18CSI7302L-R21", "DATABASE MANAGEMENT SYSTEMS", 3),
        ("U18CSI7303L-R21", "SOFTWARE ENGINEERING", 3),
    ],
    "B.TECH-ELECTRONICS AND COMMUNICATION ENGINEERING": [
        ("U18ECI7401L-R21", "DIGITAL SIGNAL PROCESSING", 4),
        ("U18ECI7402L-R21", "VLSI DESIGN", 3),
        ("U18ECI7403L-R21", "MICROPROCESSORS", 3),
    ],
    "M.E-COMPUTER SCIENCE AND ENGINEERING": [
        ("U18PGC5101L-R21", "ADVANCED ALGORITHMS", 4),
        ("U18PGC5102L-R21", "CLOUD COMPUTING", 3),
    ],
    "M.E-VLSI DESIGN": [
        ("U18PGV5101L-R21", "ADVANCED VLSI", 4),
        ("U18PGV5102L-R21", "EMBEDDED SYSTEMS", 3),
    ],
    "MBA-BUSINESS ADMINISTRATION": [
        ("U18MBA5101L-R21", "MANAGEMENT PRINCIPLES", 4),
        ("U18MBA5102L-R21", "MARKETING MANAGEMENT", 3),
    ],
    "MBA-FINANCE": [
        ("U18MBF5101L-R21", "FINANCIAL ACCOUNTING", 4),
        ("U18MBF5102L-R21", "INVESTMENT ANALYSIS", 3),
    ],
}

# Semester per program
SEMESTER = {
    "B.TECH-INFORMATION TECHNOLOGY": "EVEN UG III R21",
    "B.TECH-COMPUTER SCIENCE AND ENGINEERING": "EVEN UG III R21",
    "B.TECH-ELECTRONICS AND COMMUNICATION ENGINEERING": "EVEN UG III R21",
    "M.E-COMPUTER SCIENCE AND ENGINEERING": "EVEN PG I R21",
    "M.E-VLSI DESIGN": "EVEN PG I R21",
    "MBA-BUSINESS ADMINISTRATION": "EVEN PG I R21",
    "MBA-FINANCE": "EVEN PG I R21",
}

# Student counts per program
STUDENT_COUNT = {
    "B.TECH-INFORMATION TECHNOLOGY": 15,
    "B.TECH-COMPUTER SCIENCE AND ENGINEERING": 12,
    "B.TECH-ELECTRONICS AND COMMUNICATION ENGINEERING": 10,
    "M.E-COMPUTER SCIENCE AND ENGINEERING": 8,
    "M.E-VLSI DESIGN": 6,
    "MBA-BUSINESS ADMINISTRATION": 10,
    "MBA-FINANCE": 8,
}

# Prefix per program
PREFIX = {
    "B.TECH-INFORMATION TECHNOLOGY": ("23BIT", "BIT"),
    "B.TECH-COMPUTER SCIENCE AND ENGINEERING": ("23CSE", "CSE"),
    "B.TECH-ELECTRONICS AND COMMUNICATION ENGINEERING": ("23ECE", "ECE"),
    "M.E-COMPUTER SCIENCE AND ENGINEERING": ("23PGC", "PGC"),
    "M.E-VLSI DESIGN": ("23PGV", "PGV"),
    "MBA-BUSINESS ADMINISTRATION": ("23MBA", "MBA"),
    "MBA-FINANCE": ("23MBF", "MBF"),
}

FIRST_NAMES = [
    "ARUN", "PRIYA", "KUMAR", "MEENA", "RAHUL", "DIVYA", "SATHISH", "KAVYA",
    "VIJAY", "ANITHA", "SURESH", "LATHA", "PRASAD", "REKHA", "GANESH", "DEEPA",
    "BHARATH", "NITHYA", "RAVI", "POOJA", "MANOJ", "SUDHA", "KARTHICK", "ARCHANA",
    "SENTHIL", "VIMALA", "MURUGAN", "SARANYA", "DINESH", "THENMOZHI", "VENKAT",
    "GOMATHI", "ABISHEK", "PAVITHRA", "BALA", "MALATHI", "HARISH", "JAYA",
    "NAVEEN", "REVATHI", "AJITH", "SANGEETHA", "VIVEK", "KAMALI"
]

LAST_NAMES = [
    "M", "S", "K", "R", "P", "V", "A", "T", "B", "N", "G", "J", "L", "C"
]

STAFF = {
    "B.TECH-INFORMATION TECHNOLOGY": ["K00685-NITHYA ROOPA S", "K00340-KANAGARAJ S"],
    "B.TECH-COMPUTER SCIENCE AND ENGINEERING": ["K00421-RAMESH T", "K00512-ANITHA R"],
    "B.TECH-ELECTRONICS AND COMMUNICATION ENGINEERING": ["K00213-SURESH V", "K00314-PRAVEEN K"],
    "M.E-COMPUTER SCIENCE AND ENGINEERING": ["K00601-DIVYA M", "K00702-SENTHIL P"],
    "M.E-VLSI DESIGN": ["K00803-KUMAR J", "K00904-DEEPA S"],
    "MBA-BUSINESS ADMINISTRATION": ["K00111-VIJAY A", "K00222-MEENA T"],
    "MBA-FINANCE": ["K00333-GANESH B", "K00444-KAVYA N"],
}

random.seed(42)

# ── Build rows ────────────────────────────────────────────────────────────────

HEADERS = [
    "S.No.", "Program", "Regn. No.", "Student Name",
    "Course Code", "Course Name", "Semester", "Credits",
    "No.of hours as per credits", "No. of hours conducted", "No. of hours posted",
    "No. of hours attended", "Attendance percentage",
    "Absent", "Absent percentage",
    "Leave", "Leave percentage",
    "OnDuty", "OnDuty percentage",
    "Medical leave", "Medical leave percentage",
    "Regulation attendance percentage", "Internal Attendance percentage",
    "Not posted hours", "Staff details"
]

rows = []
sno = 1

for dept_name, programs in DEPARTMENTS.items():
    for program in programs:
        subjects = SUBJECTS[program]
        semester = SEMESTER[program]
        n_students = STUDENT_COUNT[program]
        reg_prefix, _ = PREFIX[program]
        staff = ", ".join(STAFF[program])

        # Generate unique student list for this program
        used_names = set()
        students = []
        for i in range(1, n_students + 1):
            while True:
                fname = random.choice(FIRST_NAMES)
                lname = random.choice(LAST_NAMES)
                full = f"{fname} {lname}"
                if full not in used_names:
                    used_names.add(full)
                    break
            reg = f"{reg_prefix}{i:03d}"
            students.append((reg, full))

        for reg_no, student_name in students:
            for course_code, course_name, credits in subjects:
                hours_per_credit = 30
                conducted = random.randint(18, 30)
                posted = conducted
                attended = random.randint(max(0, conducted - 8), conducted)
                absent = conducted - attended
                leave = random.randint(0, min(2, absent))
                od = random.randint(0, min(2, absent - leave))
                ml = random.randint(0, min(1, absent - leave - od))

                att_pct = round((attended / conducted) * 100) if conducted > 0 else 0
                abs_pct = round((absent / conducted) * 100) if conducted > 0 else 0
                leave_pct = round((leave / conducted) * 100) if conducted > 0 else 0
                od_pct = round((od / conducted) * 100) if conducted > 0 else 0
                ml_pct = round((ml / conducted) * 100) if conducted > 0 else 0

                effective_attended = attended + od + ml
                reg_att_pct = round((effective_attended / conducted) * 100) if conducted > 0 else 0
                internal_att_pct = att_pct
                not_posted = 0

                rows.append([
                    sno, program, reg_no, student_name,
                    course_code, course_name, semester, credits,
                    hours_per_credit, conducted, posted,
                    attended, att_pct,
                    absent, abs_pct,
                    leave, leave_pct,
                    od, od_pct,
                    ml, ml_pct,
                    reg_att_pct, internal_att_pct,
                    not_posted, staff
                ])
                sno += 1

# ── Write to Excel ────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Row 1: blank (matching original)
ws.append([None] * len(HEADERS))

# Row 2: Headers with bold styling
ws.append(HEADERS)
header_row = ws[2]
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in header_row:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

# Alternate row colours
fill_even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

for i, row_data in enumerate(rows):
    ws.append(row_data)
    excel_row = ws[i + 3]   # +2 for blank + header, +1 for 1-indexed
    fill = fill_even if i % 2 == 0 else fill_odd
    for cell in excel_row:
        cell.fill = fill

# Column widths
col_widths = [6, 40, 14, 25, 22, 35, 18, 8, 10, 10, 10, 10, 12, 8, 10,
              8, 10, 8, 10, 12, 14, 14, 14, 12, 45]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.row_dimensions[2].height = 45
ws.freeze_panes = "A3"

out_path = r"c:\Users\Lenovo\Desktop\perfoma\Dummy Multi Dept.xlsx"
wb.save(out_path)
print(f"✅ Saved: {out_path}")
print(f"   Total data rows: {len(rows)}")
print(f"   Programs: {list(SUBJECTS.keys())}")
